import openpyxl


def getRowCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column

def getCellData(path,sheetName,rowNum,colNum):   # for CellData we have to add rowNum and colNum
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value

def setCellData(path,sheetName,rowNum,colNum,data):  # for add cellData we have to add rowNum,colNum,data
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data

    workbook.save(path)

path="..//excel//testdata.xlsx"
SheetName='LoginTest'

rows = getRowCount(path,SheetName)
cols = getColCount(path,SheetName)

print(rows, "---", cols)

print(getCellData(path,SheetName,2,1))
setCellData(path,SheetName,1,4, "DOB")